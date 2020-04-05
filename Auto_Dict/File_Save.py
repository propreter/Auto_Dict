import os

from openpyxl import workbook
from openpyxl.styles import Alignment, Font

from File_Process import char_discriminator

# 设置文本格式
chn_char_style = '宋体'
eng_char_style = 'Times New Roman'
char_size = '12'
horizontal_alignment = 'left'
vertical_alignment = 'center'
wrap_Text = True
# 文档扩展名
xlsx_ext = '.xlsx'


# 设置表格格式
def style_setting(worksheet):
    col_cnt = 3
    for i in range(row_cnt):
        col_1_identifier = char_discriminator(str(worksheet.cell(i+1, column=1).value))
        if col_1_identifier is True:
            # 宋体
            worksheet.cell(i+1, column=1).font = Font(name=chn_char_style, sz=char_size)
        else:
            # Times New Roman
            worksheet.cell(i+1, column=1).font = Font(name=eng_char_style, sz=char_size)

        col_2_identifier = char_discriminator(str(worksheet.cell(i+1, column=2).value))
        if col_2_identifier is True:
            # 宋体
            worksheet.cell(i+1, column=2).font = Font(name=chn_char_style, sz=char_size)
        else:
            # Times New Roman
            worksheet.cell(i+1, column=2).font = Font(name=eng_char_style, sz=char_size)
        for j in range(col_cnt):
            worksheet.cell(i+1, j+1).alignment = Alignment(horizontal=horizontal_alignment,
                                                           vertical=vertical_alignment,
                                                           wrapText=wrap_Text
                                                           )
        worksheet.cell(i+1, column=3).font = Font(name=eng_char_style, sz=char_size)

    worksheet.column_dimensions['A'].width = 40.0
    worksheet.column_dimensions['B'].width = 40.0
    worksheet.column_dimensions['C'].width = 20.0


# 创建词汇表
def create_excel(sheet_name):
    global wb, ws
    wb = workbook.Workbook()
    ws = wb.active
    ws.title = sheet_name


# 保存释义
def save_entry_2_excel(word_dict, file_path, table_name, file_name):
    global row_cnt
    i = 0
    row_cnt = len(word_dict)
    create_excel(table_name)
    for n, v in word_dict.items():
        p = v.split('\\', 1)[0]
        e = v.split('\\', 1)[1:]
        ws.cell(row=i + 1, column=1, value=n)
        ws.cell(row=i+1, column=2, value=e[0].replace('\\[', '['))
        ws.cell(row=i+1, column=3, value=p)
        i += 1
        print('\r保存进度： {:.2f}%'.format(i * 100 / row_cnt), end='')
    save2excel(file_path, file_name)


# 保存文本
def save_text_2_excel(word_dict, file_path, table_name, file_name):
    global row_cnt
    i = 0
    row_cnt = len(word_dict)
    create_excel(table_name)
    for n, v in word_dict.items():
        ws.cell(row=i + 1, column=1, value=n)
        ws.cell(row=i + 1, column=2, value=v)
        i += 1
        print('\r保存进度： {:.2f}%'.format(i * 100 / row_cnt), end='')
    save2excel(file_path, file_name)


def save2excel(path, name):
    style_setting(ws)
    file = os.path.join(path, name)
    wb.save(file + xlsx_ext)
